from openpyxl import *
from tkinter import *
import time
from PIL import Image, ImageTk

wb = load_workbook('C:\\Users\\Muhammad Usama Bhatt\\Desktop\\sheet.xlsx')
sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 25

    sheet.cell(row=1, column=1).value = "Visitors Name"
    sheet.cell(row=1, column=2).value = "Company"
    sheet.cell(row=1, column=3).value = "Purpose"
    sheet.cell(row=1, column=4).value = "Meeting With"
    sheet.cell(row=1, column=5).value = "Date and Time"

def focus1(event):
    company_field.focus_set()

def focus2(event):
    purpose_field.focus_set()

def focus3(event):
    meeting_with_field.focus_set()

def focus4(event):
    date_time_field.focus_set()

def clear():
    name_field.delete(0, END)
    company_field.delete(0, END)
    purpose_field.delete(0, END)
    meeting_with_field.delete(0, END)
    date_time_field.delete(0, END)
    meeting_options.set('Select')
    departments_options.set('Select')

def display_time():
    date_time_field.get()
    if date_time_field.get() != "":
        date_time_field.delete(0, END)
    date_time_field.insert(0, time.asctime())

def insert():
    if (name_field.get() == "" and
            company_field.get() == "" and
            purpose_field.get() == "" and
            meeting_with_field.get() == "" and
            date_time_field.get() == ""):
        print("empty input")

    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = company_field.get()
        sheet.cell(row=current_row + 1, column=3).value = purpose_field.get()
        sheet.cell(row=current_row + 1, column=4).value = meeting_with_field.get()
        sheet.cell(row=current_row + 1, column=5).value = date_time_field.get()

        wb.save('C:\\Users\\Muhammad Usama Bhatt\\Desktop\\sheet.xlsx')

        name_field.focus_set()
        clear()

def select_purpose(choice):
    purpose_field.delete(0, END)
    purpose_field.insert(0, choice)

def show_department(choice):
    meeting_with_field.delete(0, END)
    meeting_with_field.insert(0, choice)


if __name__ == "__main__":
    root = Tk()
    root.title("Visitors Registration Form J7 Icon")
    root.iconbitmap("C:\\Users\\Muhammad Usama Bhatt\\Desktop\\j7-favicon.ico")
    root.geometry("1000x600")
    root.configure(bg='white')

    # Read the Image
    image = Image.open("C:\\Users\\Muhammad Usama Bhatt\\Desktop\\j7-favicon.png")

    # Resize the image using resize() method
    resize_image = image.resize((120, 120))

    img = ImageTk.PhotoImage(resize_image)

    # create label and add resize image
    logo = Label(image=img, background='white')
    logo.image = img
    logo.place(x=20, y=10)

    excel()

    heading = Label(root, text="Visitors Record", width=20, font=("bold", 35), bg='white')
    heading.place(x=240, y=40)

    name = Label(root, text="Visitor Name", width=20, font=("bold", 15), bg='white')
    name.place(x=230, y=150)

    company = Label(root, text="Company", width=18, font=("bold", 15), bg='white')
    company.place(x=230, y=200)

    purpose = Label(root, text="Purpose", width=17, font=("bold", 15), bg='white')
    purpose.place(x=230, y=250)

    purpose_options=[ 'Investment' ,'Private' , 'Business', 'Installment', 'Meeting' ,'Information', 'Other']
    meeting_options = StringVar()
    droplist=OptionMenu(root, meeting_options, *purpose_options, command=select_purpose)
    droplist.config(width=10)
    meeting_options.set('Select')
    droplist.place(x=730,y=250)


    meeting_with = Label(root, text="Meeting With", width=21, font=("bold", 15), bg='white')
    meeting_with.place(x=230, y=300)

    departments=[ 'Chairman', 'CEO', 'PD' ,'EDO' , 'HR' ,'IT' ,'Finance', 'Recovery', 'GM', 'Affiliate', 'Sales', 'Media']
    departments_options = StringVar()
    droplist=OptionMenu(root, departments_options, *departments, command=show_department)
    droplist.config(width=10)
    departments_options.set('Select')
    droplist.place(x=730,y=303)


    date_time = Label(root, text="Current Time", width=21, font=('bold', 15), bg='white')
    date_time.place(x=230, y=350)

    usama = Label(root, text="Created with    ❤️by Muhammad Usama Bhatti", width=37, font=('bold', 10), bg='white')
    usama.place(x=670, y=570)

    version = Label(root, text="V0.1", width=10, font=("bold", 10), bg='white')
    version.place(x=5, y=570)

    name_field = Entry(root, width=20, font=(15))
    name_field.configure(bg='light yellow')
    name_field.place(x=500, y=155)

    company_field = Entry(root, width=20, font=(15))
    company_field.configure(bg='light yellow')
    company_field.place(x=500, y=205)

    purpose_field = Entry(root, width=20, font=(15))
    purpose_field.configure(bg='light yellow')
    purpose_field.place(x=500, y=255)

    meeting_with_field = Entry(root, width=20, font=(15))
    meeting_with_field.configure(bg='light yellow')
    meeting_with_field.place(x=500, y=305)

    date_time_field = Entry(root, width=20, font=(15))
    date_time_field.configure(bg='light yellow')
    date_time_field.place(x=500, y=355)

    name_field.bind("<Return>", focus1)

    company_field.bind("<Return>", focus2)

    purpose_field.bind("<Return>", focus3)

    meeting_with_field.bind("<Return>", focus4)
    excel()

    getTime = Button(root, text='Get Time', width=20, bg="black", fg='white', command=display_time).place(x=330, y=430)
    sumbit = Button(root, text='Submit', width=20, bg="black", fg='white', command=insert).place(x=520, y=430)

    root.mainloop()
